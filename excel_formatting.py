"""
Created on Thu Apr 30 11:47:58 2020.

@author: Vasil
"""
from openpyxl.styles import Color, PatternFill, Font, Border
import openpyxl as opx
import re
import numpy as np
import pandas as pd
from pathlib import Path
import os
import sys

import zipfile
try:
    import zlib
    compression = zipfile.ZIP_DEFLATED
except (ImportError, AttributeError):
    compression = zipfile.ZIP_STORED


from copy import copy, deepcopy

current_os_is_win = sys.platform in ['win32', 'cygwin'] and os.name == "nt"
if current_os_is_win:
    import win32com.client

# In[]:


def search_and_colorise(work_sheet, searched_texts_list, color_num=4):
    if type(searched_texts_list) is str:
        raise Exception('list of str expected!')
    for seached in searched_texts_list:
        work_sheet.Cells.Find(seached).Interior.ColorIndex = color_num


def excel_file_formatting(file_to_format):
    """Format excell file for snapshot report.
    (variant for module "win32com.client")
    Parameters
    ----------
    file_to_format : str | Path
        Path of the file to e formatted.

    Returns
    -------
    None.

    """

    if not current_os_is_win:
        o_excel_file_formatting(file_to_format)
        return

    semi_lite_gray = 15
    # lite_violet = 17
    # lite_orange = 40

    Excel = win32com.client.DispatchEx("Excel.Application")
    wb = Excel.Workbooks.Open(file_to_format)
    ws_rec = wb.Worksheets("Sheet01")

    search_and_colorise(ws_rec, ("01_", '10_', '20_'))
    search_and_colorise(ws_rec,
                        ('10__lost_to_compare', '10__lost_check',
                         '20__damaged_check', '20__damaged_to_compare'),
                        color_num=semi_lite_gray)  # TODO: 21/11/11

    ws_rec.Rows('1:1').WrapText = True
    ws_rec.Columns('B:AZ').ColumnWidth = 9.71
    ws_rec.Rows("1:1").EntireRow.AutoFit()

    ws_rec.Range("B2").Select()
    Excel.ActiveWindow.FreezePanes = True

    ws_pivot_adj = wb.Worksheets("Pivot_ADJ")
    ws_pivot_adj.Activate()

    search_and_colorise(ws_pivot_adj, ["Z__SELLABLE"])

    ws_pivot_adj.Columns("D:E").NumberFormat = "# ##0"

    ws_pivot_adj.Columns('A:Z').AutoFit()
    ws_pivot_adj.Range('A1:Z1').WrapText = True
    ws_pivot_adj.Columns('G:Z').ColumnWidth = 13.14
    ws_pivot_adj.Columns('C:C').ColumnWidth = 11.57
    ws_pivot_adj.Rows("1:1").EntireRow.AutoFit()
    ws_pivot_adj.Range("D2").Select()
    Excel.ActiveWindow.FreezePanes = True

    wb.Save()
    wb.Close()
    Excel.Application.Quit()
    Excel.Quit()


# %%  24


def o_search_and_colorise(work_sheet, searched_texts_list, color='EE1111'):
    if type(searched_texts_list) is str:
        raise Exception('list of str expected!')
    # openpx style
    fill_color = PatternFill(start_color=color,
                             end_color=color,
                             fill_type='solid')
    rows_, cols_ = work_sheet.max_row, work_sheet.max_column
    for searched in searched_texts_list:
        for j in range(1, cols_ + 1):
            print(work_sheet.cell(row=1, column=j).value)
            if work_sheet.cell(row=1, column=j).value:
                if work_sheet.cell(row=1, column=j).value.find(searched) >= 0:
                    work_sheet.cell(row=1, column=j).fill = fill_color


def o_load(file_xl):
    import datetime
    print(datetime.datetime.now())
    wb = opx.load_workbook(file_xl)
    print(datetime.datetime.now())
    return wb


def o_save(wb, file_xl):
    import datetime
    print(datetime.datetime.now())
    wb.save(file_xl)
    print(datetime.datetime.now())
    # wb.close()
    return None


def o_header_wrap_and_size(ws, cols_, size):
    '''
    ws_rec.Rows('1:1').WrapText = True
    ws_rec.Columns('B:AZ').ColumnWidth = 9.71
    ws_rec.Rows("1:1").EntireRow.AutoFit()
    '''
    for j in range(1, cols_ + 1):
        ws.cell(row=1, column=j).alignment = \
            opx.styles.Alignment(
            horizontal='center',  # 'general',
            vertical='bottom',
            text_rotation=0,
            wrap_text=True,
            shrink_to_fit=False,
            indent=0)
        ws.column_dimensions[opx.utils.cell.get_column_letter(j)].width = size


def o_excel_file_formatting(file_to_format):
    """Format excell file for snapshot report.
    (variant for module "openpyxl")
    Parameters
    ----------
    file_to_format : str | Path
        Path of the file to e formatted.

    Returns
    -------
    None.

    """

    fill_color = "CCFFFF"
    semi_lite_gray = "808080"

    wb = o_load(file_to_format)
    ws_rec = wb["Sheet01"]
    # ws_rec.Range("B2").Select()
    # Excel.ActiveWindow.FreezePanes = True
    ws_rec.freeze_panes = ws_rec["B2"]
    rows_, cols_ = ws_rec.max_row, ws_rec.max_column
    o_search_and_colorise(ws_rec, ("01_", '10_', '20_'), color=fill_color)
    o_search_and_colorise(ws_rec,
                          ('10__lost_to_compare', '10__lost_check',
                           '20__damaged_check', '20__damaged_to_compare'),
                          color=semi_lite_gray)  # TODO: 21/11/11

    o_header_wrap_and_size(ws_rec, cols_, 9.71)
    ws_rec.column_dimensions["A"].width = 13.71
    ws_rec.row_dimensions[1].height = 30

    ws_pivot_adj = wb["Pivot_ADJ"]
    rows_, cols_ = ws_pivot_adj.max_row, ws_pivot_adj.max_column
    ws_pivot_adj.freeze_panes = ws_pivot_adj["D2"]
    o_search_and_colorise(ws_pivot_adj, ["Z__SELLABLE"])

    #  ws_pivot_adj.Columns("D:E").NumberFormat = "# ##0"
    D = opx.utils.cell.column_index_from_string("D")
    E = opx.utils.cell.column_index_from_string("E")
    for i in range(2, rows_+1):
        ws_pivot_adj.cell(row=i, column=D).number_format = '# ### ##0'
        ws_pivot_adj.cell(row=i, column=E).number_format = '# ### ##0'
    # ws_pivot_adj.Columns('G:Z').ColumnWidth = 13.14
    # ws_pivot_adj.Range('A1:Z1').WrapText = True
    o_header_wrap_and_size(ws_pivot_adj, cols_, 13.14)
    '''
    ws_pivot_adj.Columns('A:Z').AutoFit()
    ws_pivot_adj.Columns('C:C').ColumnWidth = 11.57
    ws_pivot_adj.Rows("1:1").EntireRow.AutoFit()
    '''
    ws_pivot_adj.column_dimensions["A"].width = 13.71
    ws_pivot_adj.column_dimensions["B"].width = 23.86
    ws_pivot_adj.column_dimensions["C"].width = 11.57
    ws_pivot_adj.column_dimensions["D"].width = 18.71
    ws_pivot_adj.column_dimensions["E"].width = 18.71
    ws_pivot_adj.column_dimensions["F"].width = 6.37

    ws_pivot_adj.row_dimensions[1].height = 30

    o_save(wb, file_to_format)
    return


def o_fee_excel_file_formatting(file_to_format):

    fill_color = "CCFFFF"
    gray_color = "999999"
    lite_gray_color = "cccccc"
    semi_lite_gray = "808080"

    wb = o_load(file_to_format)

    ws = wb['fee_preview']

    o_search_and_colorise(ws, ["FEE_assumed", "coincided"])

    fee_sum_cols = [
        "estimated_variable_closing_fee",
        "estimated_order_handling_fee_per_order",
        "estimated_pick_pack_fee_per_unit",
        "estimated_weight_handling_fee_per_unit",
        "expected_fulfillment_fee_per_unit",
    ]
    o_search_and_colorise(ws, fee_sum_cols, color=gray_color)

    o_header_wrap_and_size(ws, 35, 8)
    ws.freeze_panes = "E2"

    ws.column_dimensions["A"].width = 11.0
    ws.column_dimensions["B"].width = 16.0
    ws.column_dimensions["C"].width = 1.0
    ws.column_dimensions["D"].width = 80.0
    ws.column_dimensions["E"].width = 17.86
    ws.column_dimensions["F"].width = 8.71
    ws.column_dimensions["G"].width = 1.0
    ws.column_dimensions["H"].width = 7.0
    ws.column_dimensions["I"].width = 8.71
    ws.column_dimensions["J"].width = 8.71
    ws.column_dimensions["K"].width = 8.71
    ws.column_dimensions["L"].width = 10.0
    ws.column_dimensions["M"].width = 7.0
    ws.column_dimensions["N"].width = 7.0
    ws.column_dimensions["O"].width = 7.0
    ws.column_dimensions["P"].width = 12.8

    ws.row_dimensions[1].height = 30

# =============================================================================
# #   fee differense sheet
# =============================================================================
    ws = wb['fee_diff']

    o_search_and_colorise(ws, ["FEE_assumed", "coincided"])
    o_search_and_colorise(ws, fee_sum_cols, color=lite_gray_color)

    o_header_wrap_and_size(ws, 24, 8)
    ws.freeze_panes = "E2"

    ws.column_dimensions["K"].width = 20

    ws.row_dimensions[1].height = 30

    o_save(wb, file_to_format)
    return file_to_format


def fee_excel_file_formatting(files_dict, compression=compression):

    #  o_fee_excel_file_formatting(file_to_format
    if not current_os_is_win:
        file_formating = o_fee_excel_file_formatting
        return

    file_formating = o_fee_excel_file_formatting

    xl_files_path_dict = {"old_rooles":  None,   "new_rooles":  None}

    my_response = {             # dict_to_return
        "exit_is_ok": False,
        "exit_message": "",
        "xlsx_files": {
            'USA': copy(xl_files_path_dict),
            # 'EU': copy(xl_files_path_dict),
        },
        'zip_path': None,
    }

    for country in files_dict.keys():   # TODO:  'file_formating' have to be func from 'country'
        for rooles, file in files_dict[country].items():
            my_response["xlsx_files"][country][rooles] = file_formating(file)

    # collect all formatted files into one 'zip' file !!

    parent_name = ''
    for country in files_dict.keys():   # TODO:  'file_formating' have to be fun from 'country'
        for rooles, file in files_dict[country].items():
            if not parent_name:   # creating ZIP archive with  1-st file
                parent = files_dict[country][rooles].parent
                parent_name = files_dict[country][rooles].parent.name
                zip_file = zipfile.ZipFile(parent / (parent_name + '.zip'),
                                           mode='w', compression=compression)

            zip_file.write(files_dict[country][rooles], arcname=files_dict[country][rooles].name)
            # my_response["xlsx_files"][country][rooles] = file_formating(file)

    zip_file.close()
    zip_file_path = Path(str(zipfile.Path(zip_file)))
    my_response['zip_path'] = zip_file_path  # zipfile.Path(zip_file, (parent_name + '.zip'))
    my_response["exit_is_ok"] = True
    return my_response


#  =====
if __name__ == "__main__":
    file_to_format = r'D:\OneDrive\PyCodes\SHEDULER\Premier\Premier_FEE_to_upwork.xlsx'
    o_fee_excel_file_formatting(file_to_format)
